import pandas as pd
from datetime import timedelta
import streamlit as st
import io
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_timedelta(td):
    if pd.isnull(td):
        return "00:00"
    total_minutes = int(td.total_seconds() // 60)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours:02}:{minutes:02}"

def generate_report(file):
    df = pd.read_excel(file)
    df['Inspector'] = df['שם פרטי מבקר'] + ' ' + df['שם משפחה מבקר']
    df['start_dt'] = pd.to_datetime(df['תאריך יצירה'] + ' ' + df['שעת יצירה'], dayfirst=True)
    df['end_dt'] = pd.to_datetime(df['תאריך סיום ביקורת'] + ' ' + df['שעת סיום ביקורת'], dayfirst=True)
    df['Date'] = df['start_dt'].dt.date
    df['Month'] = df['start_dt'].dt.to_period('M')
    df['Check Duration'] = df['end_dt'] - df['start_dt']

    df.sort_values(by=['Inspector', 'Date', 'start_dt'], inplace=True)
    df['Previous_End'] = df.groupby(['Inspector', 'Date'])['end_dt'].shift()
    df['Gap'] = df['start_dt'] - df['Previous_End']
    df['Wasted Time'] = df['Gap'].apply(lambda x: x if pd.notnull(x) and x > timedelta(minutes=30) else timedelta(0))

    pivot_daily_lines = df.pivot_table(index=['Inspector', 'Date'], columns='מספר קו', aggfunc='size', fill_value=0)
    pivot_daily_lines.columns = [f"קו {col}" for col in pivot_daily_lines.columns]
    pivot_daily_lines.reset_index(inplace=True)

    summary = df.groupby(['Inspector', 'Date']).agg(
        First_Check=('start_dt', 'min'),
        Last_Check=('end_dt', 'max'),
        Num_Checks=('start_dt', 'count'),
        Total_Passengers=('סה"כ נוסעים שנבדקו', 'sum'),
        Max_Check_Duration=('Check Duration', 'max'),
        Total_Wasted_Time=('Wasted Time', 'sum')
    ).reset_index()

    summary['Working_Hours'] = summary['Last_Check'] - summary['First_Check']
    summary['Working_Hours_td'] = summary['Working_Hours']
    summary['Working_Hours'] = summary['Working_Hours'].apply(format_timedelta)
    summary['Total_Wasted_Time_td'] = summary['Total_Wasted_Time']
    summary['Total_Wasted_Time'] = summary['Total_Wasted_Time'].apply(format_timedelta)
    summary['Max_Check_Duration'] = summary['Max_Check_Duration'].apply(format_timedelta)

    daily_combined = pd.merge(summary, pivot_daily_lines, on=['Inspector', 'Date'], how='left')
    daily_combined['Month'] = pd.to_datetime(daily_combined['Date']).dt.to_period('M')

    df['Month'] = df['start_dt'].dt.to_period('M')
    pivot_monthly_lines = df.pivot_table(index=['Inspector', 'Month'], columns='מספר קו', aggfunc='size', fill_value=0)
    pivot_monthly_lines.columns = [f"קו {col}" for col in pivot_monthly_lines.columns]
    pivot_monthly_lines.reset_index(inplace=True)

    monthly_base = daily_combined.copy()
    monthly_summary = monthly_base.groupby(['Inspector', 'Month']).agg(
        Total_Working_Hours_td=('Working_Hours_td', 'sum'),
        Total_Wasted_Time_td=('Total_Wasted_Time_td', 'sum'),
        Total_Passengers=('Total_Passengers', 'sum'),
        Total_Checks=('Num_Checks', 'sum')
    ).reset_index()

    monthly_summary['Effective_Hours_td'] = monthly_summary['Total_Working_Hours_td'] - monthly_summary['Total_Wasted_Time_td']
    monthly_summary['Effective_Hours'] = monthly_summary['Effective_Hours_td'].apply(format_timedelta)
    monthly_summary['Total_Working_Hours'] = monthly_summary['Total_Working_Hours_td'].apply(format_timedelta)
    monthly_summary['Total_Wasted_Time'] = monthly_summary['Total_Wasted_Time_td'].apply(format_timedelta)
    monthly_summary['Efficiency_%'] = (monthly_summary['Effective_Hours_td'] / monthly_summary['Total_Working_Hours_td'] * 100).round(2)

    total_all_work = monthly_summary['Total_Working_Hours_td'].sum()
    total_all_effective = monthly_summary['Effective_Hours_td'].sum()

    monthly_summary['% of Working'] = (monthly_summary['Total_Working_Hours_td'] / total_all_work * 100).round(2)
    monthly_summary['% of Effective'] = (monthly_summary['Effective_Hours_td'] / total_all_effective * 100).round(2)

    monthly_summary = monthly_summary.drop(columns=['Total_Working_Hours_td', 'Total_Wasted_Time_td', 'Effective_Hours_td'])
    monthly_final = pd.merge(monthly_summary, pivot_monthly_lines, on=['Inspector', 'Month'], how='left')

    total_per_line = pivot_monthly_lines.drop(columns=['Inspector', 'Month']).sum().to_frame(name='Total Checks')
    total_per_line.reset_index(inplace=True)
    total_per_line.rename(columns={'index': 'קו'}, inplace=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for inspector, group in daily_combined.groupby('Inspector'):
            employee_sheet = group.copy()
            sums = employee_sheet.drop(columns=['First_Check','Last_Check','Month']).select_dtypes(include='number').sum(numeric_only=True)
            sums_row = pd.DataFrame([sums], index=['المجموع'])
            employee_sheet = pd.concat([employee_sheet, sums_row], ignore_index=False)
            sheet_name = inspector[:30]
            employee_sheet.drop(columns=['Working_Hours_td','Total_Wasted_Time_td'], errors='ignore').to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for sheet_name, df_sheet in zip(["Daily Summary", "Monthly Summary", "Total Per Line"], [daily_combined, monthly_final, total_per_line]):
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    output.seek(0)
    return output

st.set_page_config(page_title="تقرير فحوصات الباصات")
st.title("📋 توليد تقرير فحوصات المفتشين")

uploaded_file = st.file_uploader("📂 اختر ملف Excel للفحوصات:", type=["xlsx"])

if uploaded_file:
    output_file = generate_report(uploaded_file)
    st.success("✅ تم توليد التقرير بنجاح")
    st.download_button(
        label="📥 تحميل التقرير النهائي",
        data=output_file,
        file_name="דוח_המבקרים.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )